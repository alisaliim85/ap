import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.utils.translation import gettext_lazy as _

def generate_empty_template():
    """
    Generates an empty Excel template for member bulk upload.
    Returns a BytesIO object containing the workbook.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members Template"
    
    # Define Headers
    headers = [
        "National ID / Iqama (Required)",
        "Full Name (Required)",
        "Mobile Number (Required)",
        "Birth Date (YYYY-MM-DD) (Required)",
        "Gender (M/F) (Required)",
        "Relationship (EMPLOYEE/SPOUSE/CHILD) (Required)",
        "Sponsor National ID (If Dependent)",
        "Policy Class (Optional)",
        "Medical Card ID (Optional)",
        "National Address (Optional)",
    ]
    
    # Minimal validation/instruction helpers in the second row (optional, 
    # but good for "smart" UX, maybe just comments or styling)
    
    # Write Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        
        # Style Header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal-600
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column width
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # Freeze header
    ws.freeze_panes = "A2"
    
    # Create the stream
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

from .models import Member
from clients.models import Client
from policies.models import PolicyClass
from django.db import transaction

def process_bulk_upload(file, client):
    """
    Parses the uploaded Excel file and processes members.
    Returns a dict with 'success' list and 'failed' list.
    """
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    results = {
        'success': [],
        'failed': [],
        'total_rows': 0
    }
    
    # 1. Read all rows to avoid processing duplicates within the file itself
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    results['total_rows'] = len(rows)
    
    processed_nids = set() # To track duplicates inside the file
    
    # national_id is now a normal CharField, so we can query the DB directly
    member_cache = {}

    # Validation loop
    for index, row in enumerate(rows, start=2):
        # Unpack row (handle potentially empty rows or extra columns carefully)
        try:
            # Skip empty rows
            if not any(row):
                continue

            nid = str(row[0]).strip() if row[0] else None
            full_name = str(row[1]).strip() if row[1] else None
            mobile = str(row[2]).strip() if row[2] else None
            birth_date = row[3] # Should be datetime or date
            gender = str(row[4]).strip().upper() if row[4] else None
            relation = str(row[5]).strip().upper() if row[5] else None
            sponsor_nid = str(row[6]).strip() if row[6] else None
            policy_class_name = str(row[7]).strip() if row[7] else None # Optional
            medical_card = str(row[8]).strip() if row[8] else None # Optional
            address = str(row[9]).strip() if row[9] else "" # Optional
        except Exception as e:
            results['failed'].append({
                'row': index, 
                'name': 'Unknown', 
                'error': f"Formatting Error: {str(e)}"
            })
            continue

        # Basic Required Checks
        if not nid or not full_name or not mobile or not birth_date or not gender or not relation:
            results['failed'].append({'row': index, 'name': full_name or 'Unknown', 'error': "Missing required fields"})
            continue
            
        # Check in-file duplicate
        if nid in processed_nids:
             results['failed'].append({'row': index, 'name': full_name, 'error': "Duplicate National ID in file"})
             continue
        processed_nids.add(nid)

        # Check in-DB duplicate via direct query (national_id is now a normal field)
        existing_member = member_cache.get(nid) or Member.objects.select_related('client').filter(national_id=nid).first()
        if existing_member:
            # Check context
            if existing_member.client == client:
                msg = "Member already exists in this company"
            else:
                msg = f"Member exists in another company: {existing_member.client.name_en}"
            
            results['failed'].append({'row': index, 'name': full_name, 'error': msg, 'nid': nid})
            continue
            
        # Logic for Sponsor / Policy Class
        sponsor_obj = None
        target_policy_class = None
        
        # Resolve Sponsor if dependent
        if relation != 'PRINCIPAL' and relation != 'EMPLOYEE':
            if not sponsor_nid:
                results['failed'].append({'row': index, 'name': full_name, 'error': "Sponsor ID required for dependents"})
                continue
                
            # Look for sponsor via cache or DB query
            sponsor_obj = member_cache.get(sponsor_nid) or Member.objects.select_related('client', 'policy_class').filter(national_id=sponsor_nid, client=client).first()
            
            # Additional check: Sponsor must belong to the SAME client
            if not sponsor_obj or sponsor_obj.client != client:
                 results['failed'].append({'row': index, 'name': full_name, 'error': f"Sponsor not found in this client (ID: {sponsor_nid})"})
                 continue
            
            # Policy Class Check - Must match Sponsor
            sponsor_class = sponsor_obj.policy_class
            target_policy_class = sponsor_class
            
            if policy_class_name:
                # If provided, verify strict match
                if policy_class_name.lower() != sponsor_class.name.lower(): 
                     results['failed'].append({
                         'row': index, 
                         'name': full_name, 
                         'error': f"Policy Class Mismatch: Sponsor is '{sponsor_class.name}', provided '{policy_class_name}'"
                     })
                     continue
        else:
            # RELATION == PRINCIPAL or EMPLOYEE
            if not policy_class_name:
                 results['failed'].append({'row': index, 'name': full_name, 'error': "Policy Class required for Principal/Employee"})
                 continue
            
            # Find class by Name within Client context
            # Try direct match
            candidates = PolicyClass.objects.filter(
                name__iexact=policy_class_name, 
                policy__client=client
            )
            # Try parent company policies if applicable (Shared policies)
            if not candidates.exists() and client.parent:
                 candidates = PolicyClass.objects.filter(
                    name__iexact=policy_class_name, 
                    policy__client=client.parent,
                    policy__master_policy__isnull=True
                 )
            
            if not candidates.exists():
                results['failed'].append({'row': index, 'name': full_name, 'error': f"Invalid Policy Class: {policy_class_name}"})
                continue
            
            target_policy_class = candidates.first()

        try:
            # Map Gender / Relation keys
            g_map = {'MALE': 'M', 'M': 'M', 'FEMALE': 'F', 'F': 'F'}
            r_map = {
                'PRINCIPAL': 'PRINCIPAL', 'EMPLOYEE': 'PRINCIPAL',
                'SPOUSE': 'SPOUSE', 'CHILD': 'CHILD', 
                'SON': 'CHILD', 'DAUGHTER': 'CHILD',
                'PARENT': 'PARENT', 'BROTHER': 'BROTHER', 'SISTER': 'SISTER',
                'OTHER': 'OTHER'
            }
            
            clean_gender = g_map.get(gender.upper())
            clean_relation = r_map.get(relation.upper(), 'OTHER')
            
            if not clean_gender:
                 results['failed'].append({'row': index, 'name': full_name, 'error': "Invalid Gender"})
                 continue
            
            member = Member(
                client=client,
                full_name=full_name,
                national_id=nid,
                phone_number=mobile,
                birth_date=birth_date,
                gender=clean_gender,
                relation=clean_relation,
                sponsor=sponsor_obj,
                policy_class=target_policy_class,
                medical_card_number=medical_card,
                national_address=address
            )
            member.save()
            results['success'].append({'row': index, 'name': full_name, 'nid': nid})
            
            # Update cache so subsequent rows can find this sponsor
            member_cache[nid] = member
            
        except Exception as e:
            results['failed'].append({'row': index, 'name': full_name, 'error': f"Save Error: {str(e)}"})
            continue

    return results
